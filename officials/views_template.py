from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO


@login_required
def generate_officials_template(request):
    """Generate a sample Excel template for officials import."""
    # Create a new workbook and select the active worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Officials Template"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    # Add headers
    headers = ["name", "email", "phone", "proficiency", "certification"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 25  # name
    ws.column_dimensions['B'].width = 30  # email
    ws.column_dimensions['C'].width = 15  # phone
    ws.column_dimensions['D'].width = 15  # proficiency
    ws.column_dimensions['E'].width = 20  # certification
    
    # Sample data with entries for each proficiency level
    sample_data = [
        ["John Smith", "john.smith@example.com", "555-123-4567", "Beginner", "S&T"],
        ["Jane Doe", "jane.doe@example.com", "555-987-6543", "Intermediate", "Referee"],
        ["Mike Johnson", "mike.j@example.com", "555-555-1234", "Advanced", "Starter"],
        ["Sarah Williams", "sarah.w@example.com", "555-222-3333", "Expert", "DR"],
        ["Chris Lee", "chris.l@example.com", "555-444-5555", "Provisional", ""],
    ]
    
    # Add sample data
    for row_num, row_data in enumerate(sample_data, 2):
        for col_num, cell_value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=cell_value)
            cell.border = thin_border
    
    # Create a BytesIO object to save the workbook to
    output = BytesIO()
    wb.save(output)
    
    # Create the HttpResponse with the appropriate Excel headers
    output.seek(0)
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=officials_template.xlsx'
    
    return response
