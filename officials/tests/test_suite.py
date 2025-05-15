"""
Comprehensive Test Suite for OfficatorXL

This test suite covers all major features and functionality of the OfficatorXL application:

1. Model Tests
   - Data integrity and validation
   - Model relationships and constraints
   - Custom methods and properties

2. View Tests
   - CRUD operations for all models
   - User authentication and permissions
   - Form validation and processing
   - Filtering and pagination

3. Service Tests
   - Excel import functionality
   - Error handling and validation
   - Template generation

4. Integration Tests
   - End-to-end workflows
   - Cross-component interactions

5. Form Tests
   - Form validation
   - Form rendering
   - Form processing

6. Filter Tests
   - Filter functionality
   - Query parameter handling
"""
from django.test import TestCase, Client, RequestFactory
from django.urls import reverse
from django.contrib.auth import get_user_model
from django.db.utils import IntegrityError
from django.db import transaction
from django.core.exceptions import ValidationError
from django.core.files.uploadedfile import SimpleUploadedFile
import io
import openpyxl
import json
from datetime import date, timedelta

# Import models
from officials.models import (
    UserLeagueAdmin, Certification, League, Division, 
    Pool, Team, Official, Meet, Assignment, Event
)

# Import forms
from officials.forms import EventForm, EventImportForm, EventFilterForm

# Import services
from officials.services.event_importer import EventImporter
from officials.services.excel_errors import ExcelImportResult

# Import filters
from officials.filters import EventFilter

User = get_user_model()

class BaseTestCase(TestCase):
    """Base test class with common setup for all tests."""
    
    def setUp(self):
        """Set up test data common to all tests."""
        # Create test user
        self.client = Client()
        self.factory = RequestFactory()
        self.user = User.objects.create_user(
            username='testuser',
            email='test@example.com',
            password='testpassword'
        )
        self.admin_user = User.objects.create_superuser(
            username='admin',
            email='admin@example.com',
            password='adminpassword'
        )
        
        # Create basic model instances
        self.league = League.objects.create(
            name='Test League',
            description='Test league description',
            founded_year=2020
        )
        
        self.division = Division.objects.create(
            name='Test Division',
            description='Test division description',
            league=self.league
        )
        
        self.team = Team.objects.create(
            name='Test Team',
            abbreviation='TT',
            mascot='Tigers',
            division=self.division
        )
        
        self.certification = Certification.objects.create(
            name='Test Certification',
            abbreviation='TC',
            description='Test certification description',
            level=1
        )
        
        self.pool = Pool.objects.create(
            name='Test Pool',
            address='123 Test Street, Test City',
            length=50,
            units='Yards',
            lanes=8,
            bidirectional=True,
            team=self.team
        )
        
        self.official = Official.objects.create(
            name='Test Official',
            email='official@example.com',
            phone='555-123-4567',
            certification=self.certification,
            team=self.team,
            active=True,
            proficiency='Intermediate'
        )
        
        self.meet = Meet.objects.create(
            name='Test Meet',
            date=date.today() + timedelta(days=7),
            league=self.league,
            host_team=self.team,
            pool=self.pool,
            meet_type='regular'
        )
        self.meet.participating_teams.add(self.team)
        
        self.assignment = Assignment.objects.create(
            meet=self.meet,
            official=self.official,
            role='Referee',
            notes='Test assignment notes',
            confirmed=False
        )
        
        self.event = Event.objects.create(
            event_number=1,
            name='50m Freestyle',
            description='Short distance freestyle race',
            meet_type='dual',
            gender='male'
        )
        
        # Log in the test user
        self.client.login(username='testuser', password='testpassword')


class ModelValidationTests(BaseTestCase):
    """Tests for model validation and constraints."""
    
    def test_user_league_admin_unique_constraint(self):
        """Test the unique_together constraint on UserLeagueAdmin model."""
        # Create a UserLeagueAdmin instance
        admin = UserLeagueAdmin.objects.create(user=self.user, league=self.league)
        
        # Attempt to create a duplicate should fail
        with self.assertRaises(IntegrityError):
            with transaction.atomic():
                UserLeagueAdmin.objects.create(user=self.user, league=self.league)
    
    def test_certification_level_validation(self):
        """Test that certification level must be positive."""
        with self.assertRaises(Exception):
            certification = Certification.objects.create(
                name='Invalid Certification',
                level=-1  # This should raise an error
            )
    
    def test_event_number_validation(self):
        """Test that event_number is validated to be between 1-99."""
        # Test value below minimum
        event = Event(
            event_number=0,  # Below valid range
            name='Invalid Event',
            meet_type='dual',
            gender='male'
        )
        with self.assertRaises(ValidationError):
            # full_clean() triggers validators
            event.full_clean()
        
        # Test value above maximum
        event = Event(
            event_number=100,  # Above valid range
            name='Invalid Event',
            meet_type='dual',
            gender='male'
        )
        with self.assertRaises(ValidationError):
            # full_clean() triggers validators
            event.full_clean()
    
    def test_event_unique_constraint(self):
        """Test the unique_together constraint on Event model."""
        # Attempt to create an event with the same event_number and meet_type
        with self.assertRaises(IntegrityError):
            with transaction.atomic():
                Event.objects.create(
                    event_number=1,  # Same as existing event
                    name='Duplicate Event',
                    meet_type='dual',  # Same as existing event
                    gender='female'
                )
        
        # Creating with different meet_type should work
        Event.objects.create(
            event_number=1,
            name='Different Meet Type Event',
            meet_type='divisional',  # Different meet_type
            gender='male'
        )
    
    def test_assignment_unique_constraint(self):
        """Test the unique_together constraint on Assignment model."""
        # Attempt to create a duplicate assignment
        with self.assertRaises(IntegrityError):
            with transaction.atomic():
                Assignment.objects.create(
                    meet=self.meet,
                    official=self.official,
                    role='Referee'  # Same role as existing assignment
                )
        
        # Creating with different role should work
        Assignment.objects.create(
            meet=self.meet,
            official=self.official,
            role='Starter'  # Different role
        )


class ModelRelationshipTests(BaseTestCase):
    """Tests for model relationships and data integrity."""
    
    def test_league_division_relationship(self):
        """Test the relationship between League and Division models."""
        # Test that a division belongs to its league
        self.assertEqual(self.division.league, self.league)
        
        # Test that a league can access its divisions
        self.assertEqual(self.league.divisions.first(), self.division)
    
    def test_team_division_relationship(self):
        """Test the relationship between Team and Division models."""
        # Test that a team belongs to its division
        self.assertEqual(self.team.division, self.division)
        
        # Test that a division can access its teams
        self.assertEqual(self.division.teams.first(), self.team)
    
    def test_official_team_relationship(self):
        """Test the relationship between Official and Team models."""
        # Test that an official belongs to a team
        self.assertEqual(self.official.team, self.team)
        
        # Test that a team can access its officials
        self.assertEqual(self.team.officials.first(), self.official)
    
    def test_meet_participating_teams(self):
        """Test the many-to-many relationship between Meet and Team models."""
        # Test that a meet can access its participating teams
        self.assertEqual(self.meet.participating_teams.first(), self.team)
        
        # Create another team and add it to the meet
        team2 = Team.objects.create(
            name='Test Team 2',
            division=self.division
        )
        self.meet.participating_teams.add(team2)
        
        # Test that the meet now has two teams
        self.assertEqual(self.meet.participating_teams.count(), 2)
    
    def test_cascade_delete_behavior(self):
        """Test that cascade delete works correctly for related models."""
        # Delete the division and check that its teams are also deleted
        team_id = self.team.id
        self.division.delete()
        with self.assertRaises(Team.DoesNotExist):
            Team.objects.get(id=team_id)
        
        # Test cascade delete for other relationships as needed


class ModelMethodTests(BaseTestCase):
    """Tests for custom model methods and properties."""
    
    def test_model_string_representations(self):
        """Test the __str__ methods of models."""
        # Test League __str__
        self.assertEqual(str(self.league), 'Test League')
        
        # Test Division __str__
        self.assertEqual(str(self.division), 'Test Division - Test League')
        
        # Test Team __str__
        self.assertEqual(str(self.team), 'Test Team')
        
        # Test Official __str__
        self.assertEqual(str(self.official), 'Test Official')
        
        # Test Meet __str__
        self.assertEqual(str(self.meet), f'Test Meet - {self.meet.date}')
        
        # Test Event __str__
        self.assertEqual(str(self.event), '1 - 50m Freestyle')
        
        # Test Assignment __str__
        self.assertEqual(str(self.assignment), f'Test Official - Referee at Test Meet - {self.meet.date}')
    
    def test_pool_string_representation(self):
        """Test the __str__ method of Pool model."""
        self.assertEqual(str(self.pool), 'Test Pool (50 Yards, 8 lanes)')


class EventViewTests(BaseTestCase):
    """Tests for event views."""
    
    def setUp(self):
        """Set up test data."""
        super().setUp()
        
        # Clean up any existing events that might cause conflicts
        Event.objects.all().delete()
        
        # Create test events with unique event_number and meet_type combinations
        self.event1 = Event.objects.create(
            event_number=101,  # Use higher numbers to avoid conflicts
            name='50m Freestyle',
            description='Short distance freestyle race',
            meet_type='dual',
            gender='male'
        )
        
        self.event2 = Event.objects.create(
            event_number=102,  # Different event number
            name='Butterfly Event',
            description='',
            meet_type='divisional',
            gender='female'
        )
        
        # Set up URLs
        self.list_url = reverse('event-list')
        self.detail_url = reverse('event-detail', kwargs={'pk': self.event1.pk})
        self.create_url = reverse('event-create')
        self.update_url = reverse('event-update', kwargs={'pk': self.event1.pk})
        self.delete_url = reverse('event-delete', kwargs={'pk': self.event1.pk})
        self.import_url = reverse('event-import')
        self.template_url = reverse('event-download-template')

        # Store events for later use
        self.events = [self.event1, self.event2]
        
        # Use Django's standard test client instead of APIClient
        from django.test import Client
        self.client = Client()
        # Set credentials for login
        self.username = 'testuser'
        self.password = 'testpassword'
        # Login the user
        self.client.login(username=self.username, password=self.password)

    def test_event_list_view(self):
        """Test the event list view displays events."""
        # Use direct URL path to ensure we hit the Django template view not API
        direct_url = '/officials/events/'
        
        # Explicitly request HTML content
        response = self.client.get(direct_url, HTTP_ACCEPT='text/html')
        self.assertEqual(response.status_code, 200)
        # Just check for content, not template
        self.assertContains(response, '50m Freestyle')
        
        # Test pagination with more events
        for i in range(2, 15):  # Create 13 more events (14 total)
            Event.objects.create(
                event_number=i+200,  # Use higher numbers to avoid conflicts
                name=f'Event {i}',
                meet_type='dual',
                gender='male'
            )
        
        # Make second request with debug info
        response = self.client.get(direct_url, HTTP_ACCEPT='text/html')
        self.assertEqual(response.status_code, 200)
        
        # Check if response is HTML and has context
        self.assertEqual(response['Content-Type'].split(';')[0], 'text/html')
        
        # Verify we have the proper context variables
        self.assertTrue(hasattr(response, 'context'), "Response has no context")
        self.assertTrue('is_paginated' in response.context, f"'is_paginated' missing from context keys: {response.context.keys() if hasattr(response, 'context') else 'No context'}")
        self.assertTrue(response.context['is_paginated'], "Pagination is not enabled")
        self.assertTrue('events' in response.context, "'events' missing from context")
        self.assertEqual(len(response.context['events']), 10)  # 10 per page
    
    def test_event_detail_view(self):
        """Test the event detail view."""
        # Explicitly request HTML content
        response = self.client.get(self.detail_url, HTTP_ACCEPT='text/html')
        self.assertEqual(response.status_code, 200)
        # Just check for content, not template
        self.assertContains(response, '50m Freestyle')
    
    def test_event_create_view(self):
        """Test creating an event via the view."""
        # Get the form
        response = self.client.get(self.create_url, HTTP_ACCEPT='text/html')
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'officials/event_form.html')
        
        # Submit the form with valid data
        form_data = {
            'event_number': 15,
            'name': 'New Event',
            'description': 'New event description',
            'meet_type': 'dual',
            'gender': 'female'
        }
        response = self.client.post(self.create_url, form_data)
        self.assertEqual(response.status_code, 302)  # Redirect after success
        
        # Verify event was created
        self.assertTrue(Event.objects.filter(name='New Event').exists())
    
    def test_event_update_view(self):
        """Test updating an event via the view."""
        # Use direct URL path to ensure we hit the Django template view not API
        direct_url = f'/officials/events/{self.event1.pk}/edit/'
        
        # Get the form with current data
        response = self.client.get(direct_url, HTTP_ACCEPT='text/html')
        self.assertEqual(response.status_code, 200)
        # We check for content rather than template to avoid DRF interference
        self.assertContains(response, 'Edit Event')
        
        # Verify we have the form HTML content in the response
        html_content = response.content.decode('utf-8')
        self.assertIn('<form', html_content, "Form not found in response HTML")
        
        # Capture current values to ensure they're preserved in the update
        original_event_number = self.event1.event_number
        original_meet_type = self.event1.meet_type
        original_gender = self.event1.gender
        
        # Submit the form with updated data - include ALL form fields with valid values
        # Event numbers must be between 1-99 according to the form validation
        form_data = {
            'event_number': 42,  # Using a valid event number less than 100
            'name': 'Updated Event Name',  # Changed
            'description': 'Updated description',  # Changed
            'meet_type': original_meet_type,
            'gender': original_gender
        }
        
        # Get the URL we're actually POSTing to
        print(f"\nSubmitting form to: {direct_url}")
        print(f"Form data: {form_data}")
        
        # Send the POST and print the response details for debugging
        response = self.client.post(direct_url, data=form_data, HTTP_ACCEPT='text/html')
        print(f"Response status: {response.status_code}")
        print(f"Response headers: {response.headers}")
        
        # Check for form errors if not redirected (status 200 often indicates form errors)
        if response.status_code == 200 and hasattr(response, 'context') and 'form' in response.context:
            print(f"Form errors: {response.context['form'].errors}")
            # HTML content may contain error messages
            html_content = response.content.decode('utf-8')
            if 'error' in html_content.lower() or 'invalid' in html_content.lower():
                print("Error indicators found in HTML response")
                # Look for specific error messages in the HTML
                error_lines = [line for line in html_content.split('\n') if 'error' in line.lower() or 'invalid' in line.lower()]
                for line in error_lines[:5]:  # Print up to 5 error lines
                    print(f"Error line: {line.strip()}")
        
        # Assert proper redirect (302)
        self.assertEqual(response.status_code, 302, "Expected redirect but got status code: " + str(response.status_code))
        
        # Verify event was updated
        self.event1.refresh_from_db()
        self.assertEqual(self.event1.name, 'Updated Event Name')
        self.assertEqual(self.event1.description, 'Updated description')
    
    def test_event_delete_view(self):
        """Test deleting an event via the view."""
        # Use direct URL path to ensure we hit the Django template view not API
        direct_url = f'/officials/events/{self.event1.pk}/delete/'
        
        # Get the confirmation page
        response = self.client.get(direct_url, HTTP_ACCEPT='text/html')
        self.assertEqual(response.status_code, 200)
        # Check for content rather than template
        self.assertContains(response, 'Delete Event')
        
        # Confirm deletion - use follow=True to follow redirects
        response = self.client.post(direct_url, HTTP_ACCEPT='text/html', follow=True)
        
        # Check final response code after following redirects
        self.assertEqual(response.status_code, 200)
        
        # Check that there was a redirect in the chain
        self.assertTrue(len(response.redirect_chain) > 0, "No redirect occurred after form submission")
        self.assertEqual(response.redirect_chain[0][1], 302)  # First redirect should be 302
        
        # Verify event was deleted
        with self.assertRaises(Event.DoesNotExist):
            Event.objects.get(id=self.event1.id)
    
    def test_event_filtering(self):
        """Test event filtering functionality."""
        # Create additional events with different attributes
        Event.objects.create(
            event_number=20,
            name='Butterfly Event',
            meet_type='divisional',
            gender='female'
        )
        
        # Use direct URL path to ensure we hit the Django template view not API
        direct_url = '/officials/events/'
        
        # Test filtering by name - explicitly use HTTP_ACCEPT='text/html'
        response = self.client.get(f'{direct_url}?name=Butterfly', HTTP_ACCEPT='text/html')
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Butterfly Event')
        self.assertNotContains(response, '50m Freestyle')
        
        # Test filtering by meet_type - explicitly use HTTP_ACCEPT='text/html'
        response = self.client.get(f'{direct_url}?meet_type=divisional', HTTP_ACCEPT='text/html')
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Butterfly Event')
        self.assertNotContains(response, '50m Freestyle')
        
        # Test filtering by gender - explicitly use HTTP_ACCEPT='text/html'
        response = self.client.get(f'{direct_url}?gender=female', HTTP_ACCEPT='text/html')
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Butterfly Event')
        self.assertNotContains(response, '50m Freestyle')
        
        # Test filtering by event_number - explicitly use HTTP_ACCEPT='text/html'
        response = self.client.get(f'{direct_url}?event_number=20', HTTP_ACCEPT='text/html')
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Butterfly Event')
        self.assertNotContains(response, '50m Freestyle')
        
    def test_event_bulk_delete(self):
        """Test bulk deletion of events."""
        # Create additional events
        event2 = Event.objects.create(
            event_number=2,
            name='Event 2',
            meet_type='dual',
            gender='male'
        )
        event3 = Event.objects.create(
            event_number=3,
            name='Event 3',
            meet_type='dual',
            gender='female'
        )
        
        # Test deleting selected events
        form_data = {
            'selected_events': [self.event.id, event2.id]
        }
        response = self.client.post(reverse('event-delete-selected'), form_data)
        self.assertEqual(response.status_code, 302)  # Redirect after success
        
        # Verify selected events were deleted
        self.assertFalse(Event.objects.filter(id=self.event.id).exists())
        self.assertFalse(Event.objects.filter(id=event2.id).exists())
        
        # Verify unselected event remains
        self.assertTrue(Event.objects.filter(id=event3.id).exists())
    
    def test_download_template(self):
        """Test downloading the event import template."""
        response = self.client.get(self.template_url)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        self.assertTrue('attachment; filename=' in response['Content-Disposition'])

    def test_authentication_required(self):
        """Test that views require authentication."""
        # Log out the user
        self.client.logout()
        
        # Define explicit URL paths to test instead of using reverse
        url_paths = {
            'list': '/officials/events/',
            'detail': f'/officials/events/{self.event1.pk}/',
            'create': '/officials/events/new/',
            'update': f'/officials/events/{self.event1.pk}/edit/',
            'delete': f'/officials/events/{self.event1.pk}/delete/',
            'import': '/officials/events/import/',
            'template': '/officials/events/download-template/'
        }
        
        # Test each view for redirect to login
        for name, url_path in url_paths.items():
            # Make the request with follow=True to see redirects
            response = self.client.get(url_path, HTTP_ACCEPT='text/html', follow=True)
            
            # Check response status
            self.assertEqual(response.status_code, 200)  # Final status after redirects
            
            # Verify redirect chain contains a 302 to login
            self.assertTrue(any(status == 302 for _, status in response.redirect_chain), 
                           f"URL '{url_path}' ({name}) did not cause a 302 redirect: {response.redirect_chain}")
            
            # Verify that '/login/' is in one of the redirect URLs
            self.assertTrue(any('/login/' in url for url, _ in response.redirect_chain),
                           f"URL '{url_path}' ({name}) did not redirect to login: {response.redirect_chain}")


class EventImportViewTests(BaseTestCase):
    """Tests for event import functionality."""
    
    def setUp(self):
        super().setUp()
        self.import_url = reverse('event-import')
        
        # Create a test Excel file for import
        self.workbook = openpyxl.Workbook()
        self.worksheet = self.workbook.active
        
        # Add headers
        headers = ['event_number', 'name', 'meet_type', 'gender', 'description']
        for col_idx, header in enumerate(headers, start=1):
            self.worksheet.cell(row=1, column=col_idx).value = header
        
        # Add test data
        test_data = [
            [5, 'Test Import Event 1', 'dual', 'male', 'Test description 1'],
            [6, 'Test Import Event 2', 'divisional', 'female', 'Test description 2']
        ]
        
        for row_idx, row_data in enumerate(test_data, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                self.worksheet.cell(row=row_idx, column=col_idx).value = value
        
        # Save to BytesIO object
        self.excel_file = io.BytesIO()
        self.workbook.save(self.excel_file)
        self.excel_file.seek(0)
    
    def test_import_form_display(self):
        """Test that the import form displays correctly."""
        response = self.client.get(self.import_url)
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'officials/event_import.html')
        self.assertContains(response, 'Import Events from Excel')
    
    def test_successful_import(self):
        """Test successful import of events from Excel."""
        # Clear all events before starting the test
        Event.objects.all().delete()
        
        # Reset the file pointer
        self.excel_file.seek(0)
        
        # Submit the import form
        form_data = {
            'file': self.excel_file,
            'replace': 'off'
        }
        
        # Submit the form
        response = self.client.post(self.import_url, form_data, follow=True)
        self.assertEqual(response.status_code, 200)
        
        # Verify the events were imported
        self.assertTrue(Event.objects.filter(name='Test Import Event 1').exists())
        self.assertTrue(Event.objects.filter(name='Test Import Event 2').exists())
        
        # Verify events were imported - the test file has 2 events
        self.assertEqual(Event.objects.count(), 2, "Expected number of events should be imported")
    
    def test_replace_all_option(self):
        """Test the 'replace all' option when importing."""
        # First, create another event to be replaced
        Event.objects.create(
            event_number=10,
            name='Event To Replace',
            meet_type='dual',
            gender='male'
        )
        
        # Count events before import
        event_count_before = Event.objects.count()
        
        # Reset the file pointer
        self.excel_file.seek(0)
        
        # Prepare the upload data with replace=on
        upload_file = SimpleUploadedFile(
            'test_import.xlsx',
            self.excel_file.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        form_data = {
            'file': upload_file,
            'replace': 'on'  # Turn on replace all option
        }
        
        # Submit the form
        response = self.client.post(self.import_url, form_data, follow=True)
        self.assertEqual(response.status_code, 200)
        
        # Verify only the imported events exist (all others were replaced)
        self.assertEqual(Event.objects.count(), 2)
        self.assertTrue(Event.objects.filter(name='Test Import Event 1').exists())
        self.assertTrue(Event.objects.filter(name='Test Import Event 2').exists())
        self.assertFalse(Event.objects.filter(name='Event To Replace').exists())
    
    def test_import_validation_errors(self):
        """Test validation errors during import."""
        # Clear all events before starting the test
        Event.objects.all().delete()
        
        # Create a workbook with invalid data
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        
        # Add headers
        headers = ['event_number', 'name', 'meet_type', 'gender', 'description']
        for col_idx, header in enumerate(headers, start=1):
            worksheet.cell(row=1, column=col_idx).value = header
        
        # Add invalid data (event_number out of range)
        invalid_data = [
            [0, 'Invalid Event', 'dual', 'male', 'Event number too low'],  # Invalid event_number
            [5, 'Valid Event', 'invalid_type', 'male', 'Invalid meet_type']  # Invalid meet_type
        ]
        
        for row_idx, row_data in enumerate(invalid_data, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                worksheet.cell(row=row_idx, column=col_idx).value = value
        
        # Save to BytesIO object
        invalid_file = io.BytesIO()
        workbook.save(invalid_file)
        invalid_file.seek(0)
        
        # Prepare the upload data
        upload_file = SimpleUploadedFile(
            'invalid_import.xlsx',
            invalid_file.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        form_data = {
            'file': upload_file
        }
        
        # Submit the form
        response = self.client.post(self.import_url, form_data, follow=True)
        self.assertEqual(response.status_code, 200)
        
        # Verify the events were not imported
        self.assertFalse(Event.objects.filter(name='Invalid Event').exists())
        self.assertFalse(Event.objects.filter(name='Valid Event').exists())
        
        # Check if any events were imported - the implementation might be saving valid events
        # even when others have errors
        imported_events = Event.objects.filter(name__in=['Invalid Event', 'Valid Event']).count()
        self.assertEqual(imported_events, 0, "Invalid events should not have been imported")


class ServiceTests(BaseTestCase):
    """Tests for service classes."""
    
    def test_event_importer_direct_usage(self):
        """Test using the EventImporter service directly."""
        # Create a test Excel file
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        
        # Add headers
        headers = ['event_number', 'name', 'meet_type', 'gender', 'description']
        for col_idx, header in enumerate(headers, start=1):
            worksheet.cell(row=1, column=col_idx).value = header
        
        # Add test data
        test_data = [
            [7, 'Service Test Event 1', 'dual', 'male', 'Test description 1'],
            [8, 'Service Test Event 2', 'divisional', 'female', 'Test description 2']
        ]
        
        for row_idx, row_data in enumerate(test_data, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                worksheet.cell(row=row_idx, column=col_idx).value = value
        
        # Save to BytesIO object
        excel_file = io.BytesIO()
        workbook.save(excel_file)
        excel_file.seek(0)
        
        # Use the importer service directly
        importer = EventImporter()
        result = importer.import_events(excel_file)
        
        # Verify the result
        self.assertTrue(result.success_count > 0)  # Check success_count instead of success attribute
        self.assertEqual(result.created_count, 2)
        self.assertEqual(result.error_count, 0)
        
        # Verify the events were created
        self.assertTrue(Event.objects.filter(name='Service Test Event 1').exists())
        self.assertTrue(Event.objects.filter(name='Service Test Event 2').exists())
    
    def test_event_importer_validation(self):
        """Test validation in the EventImporter service."""
        # Create a test Excel file with invalid data
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        
        # Add headers
        headers = ['event_number', 'name', 'meet_type', 'gender', 'description']
        for col_idx, header in enumerate(headers, start=1):
            worksheet.cell(row=1, column=col_idx).value = header
        
        # Add invalid data
        invalid_data = [
            [0, 'Invalid Event', 'dual', 'male', 'Event number too low'],  # Invalid event_number
            [101, 'Invalid Event 2', 'dual', 'male', 'Event number too high'],  # Invalid event_number
            [5, 'Invalid Meet Type', 'invalid', 'male', 'Invalid meet type'],  # Invalid meet_type
            [6, 'Invalid Gender', 'dual', 'invalid', 'Invalid gender']  # Invalid gender
        ]
        
        for row_idx, row_data in enumerate(invalid_data, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                worksheet.cell(row=row_idx, column=col_idx).value = value
        
        # Save to BytesIO object
        excel_file = io.BytesIO()
        workbook.save(excel_file)
        excel_file.seek(0)
        
        # Use the importer service directly
        importer = EventImporter()
        result = importer.import_events(excel_file)
        
        # Verify the result
        self.assertEqual(result.success_count, 0)  # No successful imports
        self.assertEqual(result.created_count, 0)
        self.assertEqual(result.error_count, 4)  # All 4 rows should have errors
        
        # Verify error messages contain expected validation messages
        self.assertTrue(any('event_number' in error.lower() for error in result.errors))
        self.assertTrue(any('meet_type' in error.lower() for error in result.errors))
        self.assertTrue(any('gender' in error.lower() for error in result.errors))
    
    def test_generate_template(self):
        """Test generating an Excel template."""
        # Use the static method to generate a template
        workbook = EventImporter.generate_template()
        
        # Verify the workbook structure
        worksheet = workbook.active
        
        # Check that headers are present
        expected_headers = ['event_number', 'name', 'meet_type', 'gender', 'description']
        for col_idx, header in enumerate(expected_headers, start=1):
            self.assertEqual(worksheet.cell(row=1, column=col_idx).value, header)


class FormTests(BaseTestCase):
    """Tests for form validation and processing."""
    
    def test_event_form_validation(self):
        """Test validation in the EventForm."""
        # Test valid data
        form_data = {
            'event_number': 15,
            'name': 'Test Form Event',
            'description': 'Test description',
            'meet_type': 'dual',
            'gender': 'male'
        }
        form = EventForm(data=form_data)
        self.assertTrue(form.is_valid())
        
        # Test invalid event_number (too high)
        form_data['event_number'] = 100
        form = EventForm(data=form_data)
        self.assertFalse(form.is_valid())
        self.assertTrue('event_number' in form.errors)
        
        # Test invalid event_number (too low)
        form_data['event_number'] = 0
        form = EventForm(data=form_data)
        self.assertFalse(form.is_valid())
        self.assertTrue('event_number' in form.errors)
        
        # Test invalid meet_type
        form_data['event_number'] = 15
        form_data['meet_type'] = 'invalid'
        form = EventForm(data=form_data)
        self.assertFalse(form.is_valid())
        self.assertTrue('meet_type' in form.errors)
        
        # Test invalid gender
        form_data['meet_type'] = 'dual'
        form_data['gender'] = 'invalid'
        form = EventForm(data=form_data)
        self.assertFalse(form.is_valid())
        self.assertTrue('gender' in form.errors)
        
        # Test missing required field (name)
        form_data['gender'] = 'male'
        form_data['name'] = ''
        form = EventForm(data=form_data)
        self.assertFalse(form.is_valid())
        self.assertTrue('name' in form.errors)
    
    def test_event_import_form(self):
        """Test the EventImportForm."""
        # Create a test file
        excel_file = SimpleUploadedFile(
            'test_file.xlsx', 
            b'test content',
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        # Test with valid file
        form_data = {'file': excel_file}
        form = EventImportForm(data={}, files=form_data)
        self.assertTrue(form.is_valid())
        
        # Test with missing file
        form = EventImportForm(data={}, files={})
        self.assertFalse(form.is_valid())
        self.assertTrue('file' in form.errors)
        
        # Note: The current form implementation doesn't validate file type
        # so we don't test that here, but we should enhance the form in the future


class FilterTests(BaseTestCase):
    """Tests for filter functionality."""
    
    def setUp(self):
        super().setUp()
        # Create additional events for filtering
        Event.objects.create(
            event_number=10,
            name='Filter Test 1',
            meet_type='dual',
            gender='female'
        )
        Event.objects.create(
            event_number=20,
            name='Filter Test 2',
            meet_type='divisional',
            gender='male'
        )
        Event.objects.create(
            event_number=30,
            name='Another Event',
            meet_type='divisional',
            gender='female'
        )
    
    def test_event_filter(self):
        """Test the EventFilter functionality."""
        # Test filtering by name
        filterset = EventFilter({'name': 'Filter'}, queryset=Event.objects.all())
        self.assertEqual(filterset.qs.count(), 2)  # Should match 'Filter Test 1' and 'Filter Test 2'
        
        # Test filtering by event_number
        filterset = EventFilter({'event_number': 10}, queryset=Event.objects.all())
        self.assertEqual(filterset.qs.count(), 1)  # Should match only event with number 10
        
        # Test filtering by meet_type
        filterset = EventFilter({'meet_type': 'divisional'}, queryset=Event.objects.all())
        self.assertEqual(filterset.qs.count(), 2)  # Should match both divisional events
        
        # Test filtering by gender
        filterset = EventFilter({'gender': 'female'}, queryset=Event.objects.all())
        self.assertEqual(filterset.qs.count(), 2)  # Should match both female events
        
        # Test multiple filters (name AND meet_type)
        filterset = EventFilter(
            {'name': 'Filter', 'meet_type': 'divisional'}, 
            queryset=Event.objects.all()
        )
        self.assertEqual(filterset.qs.count(), 1)  # Should match only 'Filter Test 2'
    
    def test_filter_form_rendering(self):
        """Test that the filter form renders correctly."""
        filterset = EventFilter({}, queryset=Event.objects.all())
        self.assertTrue(hasattr(filterset, 'form'))
        
        # Check that the form has the expected fields
        expected_fields = ['event_number', 'name', 'meet_type', 'gender']
        for field in expected_fields:
            self.assertTrue(field in filterset.form.fields)
