from django.test import TestCase, Client
from django.urls import reverse
from django.contrib.auth import get_user_model
from django.core.files.uploadedfile import SimpleUploadedFile
from io import BytesIO
from officials.models import Team, Division, League, Official, Certification

User = get_user_model()

class TeamViewsTest(TestCase):
    """Test cases for Team views."""
    
    def setUp(self):
        """Set up test data."""
        # Create a regular user
        self.user = User.objects.create_user(
            username='testuser', 
            email='test@example.com',
            password='testpassword123'
        )
        
        # Create a superuser (admin)
        self.admin_user = User.objects.create_superuser(
            username='adminuser',
            email='admin@example.com',
            password='adminpassword123'
        )
        
        # Create test league and associate with user
        self.league = League.objects.create(
            name='Test League',
            description='A test league',
            founded_year=2020
        )
        self.league.users.add(self.user)
        
        # Create another league not associated with user
        self.other_league = League.objects.create(
            name='Other League',
            description='Another test league',
            founded_year=2021
        )
        
        # Create divisions
        self.division = Division.objects.create(
            name='Test Division',
            description='A test division',
            league=self.league
        )
        
        self.other_division = Division.objects.create(
            name='Other Division',
            description='Another test division',
            league=self.other_league
        )
        
        # Create test teams
        self.team = Team.objects.create(
            name='Test Team',
            abbreviation='TT',
            mascot='Tigers',
            division=self.division
        )
        
        self.other_team = Team.objects.create(
            name='Other Team',
            abbreviation='OT',
            mascot='Owls',
            division=self.other_division
        )
        
        # Create certification for officials
        self.certification = Certification.objects.create(
            name='Test Certification',
            abbreviation='TC',
            description='Test certification',
            level=1
        )
        
        # Create officials
        self.official1 = Official.objects.create(
            name='Test Official 1',
            email='official1@example.com',
            phone='555-123-4567',
            certification=self.certification,
            team=self.team,
            active=True,
            proficiency='Intermediate'
        )
        
        self.official2 = Official.objects.create(
            name='Test Official 2',
            email='official2@example.com',
            phone='555-234-5678',
            certification=self.certification,
            team=self.team,
            active=False,
            proficiency='Advanced'
        )
        
        # Set up the test client
        self.client = Client()
    
    def test_team_list_view_requires_login(self):
        """Test that the team list view requires login."""
        response = self.client.get(reverse('team_list'))
        self.assertEqual(response.status_code, 302)  # Redirect to login
        self.assertIn('login', response.url)
    
    def test_team_list_view(self):
        """Test the team list view shows teams from user's leagues."""
        # Login
        self.client.login(username='testuser', password='testpassword123')
        
        # Access the team list page
        response = self.client.get(reverse('team_list'))
        
        # Check response is successful
        self.assertEqual(response.status_code, 200)
        
        # Check that only teams from user's leagues are displayed
        self.assertContains(response, 'Test Team')
        self.assertNotContains(response, 'Other Team')
        
        # Check context
        self.assertEqual(len(response.context['page_obj']), 1)
        
        # Check that filtering elements are present
        self.assertContains(response, 'name="name"')
        
        # Check card/list view toggle is present
        self.assertContains(response, 'id="view-toggle"')
    
    def test_team_detail_view(self):
        """Test the team detail view."""
        # Login
        self.client.login(username='testuser', password='testpassword123')
        
        # Access the team detail page
        response = self.client.get(reverse('team_detail', args=[self.team.id]))
        
        # Check response is successful
        self.assertEqual(response.status_code, 200)
        
        # Check that team details are displayed
        self.assertContains(response, 'Test Team')
        self.assertContains(response, 'Tigers')
        self.assertContains(response, 'Test Division')
        
        # By default, only active officials should be shown
        self.assertContains(response, 'Test Official 1')
        self.assertNotContains(response, 'Test Official 2')
        
        # Check import form is present
        self.assertContains(response, 'enctype="multipart/form-data"')
        self.assertContains(response, 'name="excel_file"')
    
    def test_team_detail_view_with_inactive_officials(self):
        """Test the team detail view with inactive officials shown."""
        # Login
        self.client.login(username='testuser', password='testpassword123')
        
        # Access the team detail page with show_inactive=true
        response = self.client.get(reverse('team_detail', args=[self.team.id]) + '?show_inactive=true')
        
        # Check response is successful
        self.assertEqual(response.status_code, 200)
        
        # Both active and inactive officials should be shown
        self.assertContains(response, 'Test Official 1')
        self.assertContains(response, 'Test Official 2')
    
    def test_team_detail_view_unauthorized(self):
        """Test that users cannot view teams they don't have access to."""
        # Login
        self.client.login(username='testuser', password='testpassword123')
        
        # Try to access unauthorized team
        response = self.client.get(reverse('team_detail', args=[self.other_team.id]))
        
        # Should redirect with error message
        self.assertEqual(response.status_code, 302)
        self.assertRedirects(response, reverse('team_list'))
    
    def test_team_create_view(self):
        """Test creating a team."""
        # Login
        self.client.login(username='testuser', password='testpassword123')
        
        # Get the create form
        response = self.client.get(reverse('team_create'))
        self.assertEqual(response.status_code, 200)
        
        # Check that only divisions from user's leagues are available
        self.assertContains(response, 'Test Division')
        self.assertNotContains(response, 'Other Division')
        
        # Post new team data
        team_data = {
            'name': 'New Test Team',
            'abbreviation': 'NTT',
            'mascot': 'Newts',
            'division': self.division.id
        }
        
        response = self.client.post(reverse('team_create'), team_data)
        
        # Should redirect to detail page of new team
        self.assertEqual(response.status_code, 302)
        
        # Check that team was created
        new_team = Team.objects.get(name='New Test Team')
        self.assertEqual(new_team.abbreviation, 'NTT')
        self.assertEqual(new_team.mascot, 'Newts')
        self.assertEqual(new_team.division, self.division)
    
    def test_team_update_view(self):
        """Test updating a team."""
        # Login
        self.client.login(username='testuser', password='testpassword123')
        
        # Get the update form
        response = self.client.get(reverse('team_update', args=[self.team.id]))
        self.assertEqual(response.status_code, 200)
        
        # Post updated team data
        updated_data = {
            'name': 'Updated Team Name',
            'abbreviation': 'UTN',
            'mascot': 'Updated Mascot',
            'division': self.division.id
        }
        
        response = self.client.post(reverse('team_update', args=[self.team.id]), updated_data)
        
        # Should redirect to detail page
        self.assertEqual(response.status_code, 302)
        self.assertRedirects(response, reverse('team_detail', args=[self.team.id]))
        
        # Check that team was updated
        updated_team = Team.objects.get(id=self.team.id)
        self.assertEqual(updated_team.name, 'Updated Team Name')
        self.assertEqual(updated_team.abbreviation, 'UTN')
        self.assertEqual(updated_team.mascot, 'Updated Mascot')
    
    def test_team_update_view_unauthorized(self):
        """Test that users cannot update teams they don't have access to."""
        # Login
        self.client.login(username='testuser', password='testpassword123')
        
        # Try to update unauthorized team
        response = self.client.get(reverse('team_update', args=[self.other_team.id]))
        
        # Should redirect with error message
        self.assertEqual(response.status_code, 302)
        self.assertRedirects(response, reverse('team_list'))
    
    def test_team_delete_view(self):
        """Test deleting a team."""
        # Login
        self.client.login(username='testuser', password='testpassword123')
        
        # Get the delete confirmation page
        response = self.client.get(reverse('team_delete', args=[self.team.id]))
        self.assertEqual(response.status_code, 200)
        
        # Confirm deletion
        response = self.client.post(reverse('team_delete', args=[self.team.id]))
        
        # Should redirect to list page
        self.assertEqual(response.status_code, 302)
        self.assertRedirects(response, reverse('team_list'))
        
        # Check that team was deleted
        self.assertFalse(Team.objects.filter(id=self.team.id).exists())
    
    def test_team_delete_view_unauthorized(self):
        """Test that users cannot delete teams they don't have access to."""
        # Login
        self.client.login(username='testuser', password='testpassword123')
        
        # Try to delete unauthorized team
        response = self.client.get(reverse('team_delete', args=[self.other_team.id]))
        
        # Should redirect with error message
        self.assertEqual(response.status_code, 302)
        self.assertRedirects(response, reverse('team_list'))
        
        # Check that team still exists
        self.assertTrue(Team.objects.filter(id=self.other_team.id).exists())
    
    def test_team_import_officials(self):
        """Test importing officials from Excel file."""
        # Login
        self.client.login(username='testuser', password='testpassword123')
        
        # Create a simple Excel file for testing
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Add headers
        headers = ['Name', 'Email', 'Phone', 'Certification', 'Proficiency']
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)
        
        # Add data
        data = [
            ['New Official', 'new@example.com', '555-999-8888', 'Test Certification', 'Advanced'],
            ['Updated Official', 'updated@example.com', '555-777-6666', 'Test Certification', 'Expert']
        ]
        
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, cell_value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=cell_value)
        
        # Save to file-like object
        file_data = BytesIO()
        wb.save(file_data)
        file_data.seek(0)
        
        # Create a SimpleUploadedFile from the Excel data
        excel_file = SimpleUploadedFile(
            name='test_import.xlsx',
            content=file_data.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        # Add an official to be updated
        Official.objects.create(
            name='Updated Official',
            email='old@example.com',  # This should be updated
            phone='555-111-2222',     # This should be updated
            certification=self.certification,
            team=self.team,
            active=True,
            proficiency='Beginner'    # This should be updated
        )
        
        # Submit the import form
        response = self.client.post(
            reverse('team_import_officials', args=[self.team.id]),
            {'excel_file': excel_file},
            follow=True
        )
        
        # Check for success message
        self.assertContains(response, 'Import successful')
        
        # Check that new official was created
        self.assertTrue(Official.objects.filter(name='New Official', team=self.team).exists())
        new_official = Official.objects.get(name='New Official', team=self.team)
        self.assertEqual(new_official.email, 'new@example.com')
        self.assertEqual(new_official.phone, '555-999-8888')
        self.assertEqual(new_official.proficiency, 'Advanced')
        
        # Check that existing official was updated
        updated_official = Official.objects.get(name='Updated Official', team=self.team)
        self.assertEqual(updated_official.email, 'updated@example.com')
        self.assertEqual(updated_official.phone, '555-777-6666')
        self.assertEqual(updated_official.proficiency, 'Expert')
        
        # Check that official2 was deactivated (since it wasn't in the import)
        official2_updated = Official.objects.get(id=self.official2.id)
        self.assertFalse(official2_updated.active)
    
    def test_team_import_officials_unauthorized(self):
        """Test that users cannot import officials for teams they don't have access to."""
        # Login
        self.client.login(username='testuser', password='testpassword123')
        
        # Create a simple Excel file
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = 'Name'
        ws['A2'] = 'Test Name'
        
        file_data = BytesIO()
        wb.save(file_data)
        file_data.seek(0)
        
        excel_file = SimpleUploadedFile(
            'test_import.xlsx',
            file_data.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        # Try to import to unauthorized team
        response = self.client.post(
            reverse('team_import_officials', args=[self.other_team.id]),
            {'file': excel_file}
        )
        
        # Should redirect with error message
        self.assertEqual(response.status_code, 302)
        self.assertRedirects(response, reverse('team_list'))
