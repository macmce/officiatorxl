from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404
from django.contrib.auth.decorators import login_required
from .models import Team, Official
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO
import json


@login_required
def export_team_officials_excel(request, pk):
    """Export team officials to Excel file."""
    team = get_object_or_404(Team, pk=pk)
    
    # Check if user has permission to view this team
    if not request.user.leagues.filter(id=team.division.league.id).exists() and not request.user.is_staff:
        return HttpResponse('Unauthorized', status=401)
    
    # Get active officials for this team
    officials = team.officials.filter(active=True)
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{team.name} Officials"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    # Add headers
    headers = ["name", "email", "phone", "proficiency", "certification"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 25  # name
    ws.column_dimensions['B'].width = 30  # email
    ws.column_dimensions['C'].width = 15  # phone
    ws.column_dimensions['D'].width = 15  # proficiency
    ws.column_dimensions['E'].width = 20  # certification
    
    # Add officials data
    for row_num, official in enumerate(officials, 2):
        # Name
        cell = ws.cell(row=row_num, column=1, value=official.name)
        cell.border = thin_border
        
        # Email
        cell = ws.cell(row=row_num, column=2, value=official.email)
        cell.border = thin_border
        
        # Phone
        cell = ws.cell(row=row_num, column=3, value=official.phone)
        cell.border = thin_border
        
        # Proficiency
        cell = ws.cell(row=row_num, column=4, value=official.proficiency)
        cell.border = thin_border
        
        # Certification
        certification_name = official.certification.name if official.certification else ""
        cell = ws.cell(row=row_num, column=5, value=certification_name)
        cell.border = thin_border
    
    # Create a BytesIO object to save the workbook to
    output = BytesIO()
    wb.save(output)
    
    # Create the HttpResponse with the appropriate Excel headers
    output.seek(0)
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={team.name}_officials.xlsx'
    
    return response


@login_required
def export_team_officials_json(request, pk):
    """Export team officials to JSON file."""
    team = get_object_or_404(Team, pk=pk)
    
    # Check if user has permission to view this team
    if not request.user.leagues.filter(id=team.division.league.id).exists() and not request.user.is_staff:
        return HttpResponse('Unauthorized', status=401)
    
    # Get active officials for this team
    officials = team.officials.filter(active=True)
    
    # Create a list of officials data
    officials_data = []
    for official in officials:
        officials_data.append({
            'name': official.name,
            'email': official.email,
            'phone': official.phone,
            'proficiency': official.proficiency,
            'certification': official.certification.name if official.certification else ""
        })
    
    # Create a JSON response
    response = HttpResponse(
        json.dumps(officials_data, indent=4),
        content_type='application/json'
    )
    response['Content-Disposition'] = f'attachment; filename={team.name}_officials.json'
    
    return response
