import openpyxl
from django.utils.text import slugify
from officials.models import Event
from officials.services.excel_errors import ExcelImportResult, ExcelValidator, ExcelHeaderError, ExcelRowError


class EventImporter(ExcelValidator):
    """Service class to handle event imports from Excel files."""
    
    REQUIRED_HEADERS = ['event_number', 'name', 'meet_type', 'gender']
    VALID_MEET_TYPES = dict(Event.MEET_TYPE_CHOICES)
    VALID_GENDERS = dict(Event.GENDER_CHOICES)
    
    def __init__(self, replace_all=False):
        """
        Initialize the importer.
        
        Args:
            replace_all: If True, all existing events will be deleted before import
        """
        self.replace_all = replace_all
        self.result = ExcelImportResult()
    
    def validate_event_number(self, value):
        """Validate event number is an integer between 1-99."""
        try:
            event_number = int(value)
            if event_number < 1 or event_number > 99:
                raise ValueError(f"Must be between 1-99, got {event_number}")
            return event_number
        except (ValueError, TypeError):
            raise ValueError(f"Must be a number between 1-99, got '{value}'")
    
    def validate_meet_type(self, value):
        """Validate meet_type is one of the valid choices."""
        meet_type = str(value).lower()
        if meet_type not in self.VALID_MEET_TYPES:
            raise ValueError(
                f"Invalid value '{meet_type}'. Must be one of: {', '.join(self.VALID_MEET_TYPES.keys())}"
            )
        return meet_type
    
    def validate_gender(self, value):
        """Validate gender is one of the valid choices."""
        gender = str(value).lower()
        if gender not in self.VALID_GENDERS:
            raise ValueError(
                f"Invalid value '{gender}'. Must be one of: {', '.join(self.VALID_GENDERS.keys())}"
            )
        return gender
    
    def validate_row(self, row_data, row_number):
        """Validate a single row of data from the Excel file."""
        # Check required fields are present
        for field in self.REQUIRED_HEADERS:
            if field not in row_data or row_data[field] is None or row_data[field] == '':
                raise ExcelRowError(row_number, f"Missing required field: {field}")
        
        # Validate field values using validators
        validators = {
            'event_number': self.validate_event_number,
            'meet_type': self.validate_meet_type,
            'gender': self.validate_gender
        }
        
        return super().validate_row(row_data, row_number, validators)
    
    def process_row(self, row_data, row_number):
        """Process a single row of data from the Excel file."""
        try:
            # Validate the row data
            self.validate_row(row_data, row_number)
            
            # Prepare data for event creation/update
            event_data = {
                'event_number': int(row_data['event_number']),
                'name': row_data['name'],
                'meet_type': row_data['meet_type'].lower(),
                'gender': row_data['gender'].lower(),
            }
            
            # Add optional description if present
            if 'description' in row_data and row_data['description']:
                event_data['description'] = row_data['description']
            
            # Try to find existing event by event_number and meet_type
            existing_event = Event.objects.filter(
                event_number=event_data['event_number'],
                meet_type=event_data['meet_type']
            ).first()
            
            if existing_event:
                # Update existing event
                for key, value in event_data.items():
                    setattr(existing_event, key, value)
                existing_event.save()
                self.result.updated_count += 1
            else:
                # Create new event
                Event.objects.create(**event_data)
                self.result.created_count += 1
                
        except ExcelRowError as e:
            # Add the error to our result object
            self.result.add_errors_from_exception(e)
    
    def import_events(self, file_obj):
        """
        Import events from an Excel file.
        
        Args:
            file_obj: An uploaded file object (e.g., from request.FILES)
            
        Returns:
            ExcelImportResult: Results of the import operation
        """
        # Reset result for new import
        self.result = ExcelImportResult()
        
        try:
            # Load workbook
            workbook = openpyxl.load_workbook(file_obj)
            worksheet = workbook.active
            
            # Get headers from first row
            headers = [cell.value.lower() if cell.value else None for cell in worksheet[1]]
            
            # Validate headers
            try:
                self.validate_headers(headers, self.REQUIRED_HEADERS)
            except ExcelHeaderError as e:
                self.result.add_errors_from_exception(e)
                return self.result
            
            # Clear existing events if replace_all is True
            if self.replace_all:
                count = Event.objects.count()
                Event.objects.all().delete()
                if count > 0:
                    self.result.skipped_count += count
            
            # Process data rows
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
                row_data = {headers[col_idx]: cell.value 
                          for col_idx, cell in enumerate(row) 
                          if col_idx < len(headers) and headers[col_idx]}
                
                # Skip empty rows
                if not any(row_data.values()):
                    self.result.skipped_count += 1
                    continue
                    
                self.process_row(row_data, row_idx)
                
        except Exception as e:
            # Catch any other exceptions and add to result
            self.result.add_error("General", f"Error processing file: {str(e)}")
            
        return self.result
    
    @staticmethod
    def generate_template():
        """Generate a template Excel file for event imports."""
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Events"
        
        # Add headers
        headers = ['event_number', 'name', 'meet_type', 'gender', 'description']
        for col_idx, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = openpyxl.styles.Font(bold=True)
        
        # Add sample data
        sample_data = [
            {'event_number': 1, 'name': '50 Free', 'meet_type': 'dual', 'gender': 'male', 'description': 'Sample description'},
            {'event_number': 2, 'name': '100 Breast', 'meet_type': 'divisional', 'gender': 'female', 'description': 'Optional'},
        ]
        
        for row_idx, data in enumerate(sample_data, start=2):
            for col_idx, header in enumerate(headers, start=1):
                worksheet.cell(row=row_idx, column=col_idx).value = data.get(header, '')
        
        # Add notes
        notes_row = len(sample_data) + 3
        worksheet.cell(row=notes_row, column=1).value = "Notes:"
        worksheet.cell(row=notes_row + 1, column=1).value = "- event_number must be between 1-99"
        worksheet.cell(row=notes_row + 2, column=1).value = f"- meet_type must be one of: {', '.join(dict(Event.MEET_TYPE_CHOICES).keys())}"
        worksheet.cell(row=notes_row + 3, column=1).value = f"- gender must be one of: {', '.join(dict(Event.GENDER_CHOICES).keys())}"
        worksheet.cell(row=notes_row + 4, column=1).value = "- description is optional"
        
        return workbook
